import csv
import io
import re
import unicodedata
import zipfile
from collections import Counter
from datetime import date, datetime, time, timedelta, timezone
from difflib import get_close_matches

from dateutil.relativedelta import relativedelta
from decimal import Decimal, InvalidOperation
from uuid import UUID
from xml.etree import ElementTree as ET

from sqlalchemy import select, tuple_
from sqlalchemy.orm import Session

from app.core.cache import invalidate_dashboard_cache
from app.models import Checkin, CheckinSource, Member, MemberStatus
from app.schemas import (
    ImportErrorEntry,
    ImportPreview,
    ImportPreviewRow,
    ImportPreviewSourceColumn,
    ImportSummary,
    MissingMemberEntry,
)
from app.services.onboarding_service import create_import_playbook_tasks_for_member
from app.services.risk import refresh_member_risk_snapshot
from app.utils.encryption import decrypt_cpf, encrypt_cpf


NAME_KEYS = ("full_name", "name", "nome", "aluno", "member_name", "cliente")
FIRST_NAME_KEYS = ("first_name", "primeiro_nome", "nome_1", "nome1")
LAST_NAME_KEYS = ("last_name", "sobrenome", "ultimo_nome", "nome_2", "nome2")
EMAIL_KEYS = ("email", "e_mail", "mail", "member_email")
PHONE_KEYS = ("phone", "telefone", "telefones", "celular", "whatsapp")
CPF_KEYS = ("cpf", "documento", "document", "cpf_cnpj")
PLAN_KEYS = (
    "plan_name",
    "plan",
    "plano",
    "plano_nome",
    "assinatura",
    "assinaturas",
    "assinatura_atual",
    "plano_atual",
    "produto",
    "produto_nome",
    "assinaturas_condicoes",
)
PLAN_NAME_SOURCE_KEYS = PLAN_KEYS[:-1]
PLAN_CONDITION_KEYS = ("assinaturas_condicoes",)
PLAN_RENEWAL_KEYS = ("dt_prox_renovacao_assinatura", "data_prox_vencimento")
MONTHLY_FEE_KEYS = ("monthly_fee", "mensalidade", "valor", "valor_mensal", "price")
JOIN_DATE_KEYS = ("join_date", "data_matricula", "data_adesao", "data_inicio", "start_date", "cadastro", "dt_primeira_ativacao", "conversao")
LAST_ACCESS_KEYS = ("ult_acesso", "ultimo_acesso", "last_access", "last_checkin_at", "dt_ultimo_login")
PREFERRED_SHIFT_KEYS = (
    "preferred_shift",
    "turno",
    "turno_preferido",
    "horario",
    "shift",
    "assinaturas_horarios",
    "horarios",
)
STATUS_KEYS = ("status", "situacao", "state")
EXTERNAL_ID_KEYS = ("external_id", "matricula", "codigo", "member_code", "id_aluno", "id_externo", "codigo_acesso")

MEMBER_ID_KEYS = ("member_id", "id_membro", "aluno_id", "member_uuid")
CHECKIN_AT_KEYS = (
    "checkin_at",
    "data_checkin",
    "checkin",
    "data_hora",
    "datetime",
    "timestamp",
    "hora",
    "data",
    "data_entrada",
    "data_hora_entrada",
)
CHECKIN_DATE_KEYS = ("checkin_date", "data_checkin", "data", "date", "data_entrada", "data_saida")
CHECKIN_TIME_KEYS = ("checkin_time", "hora_checkin", "hora", "time", "hora_entrada", "hora_saida")
CHECKIN_SOURCE_KEYS = ("source", "origem", "tipo")

DATE_FORMATS = ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y")
DATETIME_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y/%m/%d %H:%M:%S",
    "%Y/%m/%d %H:%M",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%dT%H:%M",
    "%d/%m/%Y %H:%M:%S",
    "%d/%m/%Y %H:%M",
    "%d-%m-%Y %H:%M:%S",
    "%d-%m-%Y %H:%M",
)

_MAX_MEMBER_NAME = 120
_MAX_MEMBER_EMAIL = 255
_MAX_MEMBER_PHONE = 32
_MAX_MEMBER_PLAN = 100
_MAX_MEMBER_SHIFT = 24
_NAME_PARTICLES = {"da", "de", "do", "das", "dos", "e"}
_GENERIC_PLAN_VALUES = {"livre", "nao_informado", "nao informado", "nao", "sem_assinatura", "sem plano", "sem_plano"}
_GENERIC_SHIFT_VALUES = {"livre", "nao_informado", "nao informado", "integral", "full", "all_day"}
_IGNORABLE_CHECKIN_NAMES = {"passagem liberada manualmente", "registro manual", "catraca liberada manualmente"}
_PLAN_CYCLE_LABELS = {
    "monthly": "MENSAL",
    "semiannual": "SEMESTRAL",
    "annual": "ANUAL",
}
_PLAN_CYCLE_SOURCE_UNKNOWN = "unknown"
_PLAN_CYCLE_PATTERNS = {
    "annual": re.compile(r"\b(?:12\s*mes(?:es)?|anual)\b", re.IGNORECASE),
    "semiannual": re.compile(r"\b(?:6\s*mes(?:es)?|semestral)\b", re.IGNORECASE),
    "monthly": re.compile(r"\b(?:1\s*mes(?:es)?|mensal(?:idade)?)\b", re.IGNORECASE),
}
_PLAN_CYCLE_CLEANUP_PATTERN = re.compile(
    r"\b(?:1\s*mes(?:es)?|mensal(?:idade)?|6\s*mes(?:es)?|semestral|12\s*mes(?:es)?|anual)\b",
    re.IGNORECASE,
)


_IMPORT_BATCH_SIZE = 500
_IMPORT_PREVIEW_SAMPLE_LIMIT = 5
_IMPORT_ONBOARDING_WINDOW_DAYS = 30
_MEMBER_PREVIEW_COLUMNS = set(
    NAME_KEYS
    + FIRST_NAME_KEYS
    + LAST_NAME_KEYS
    + EMAIL_KEYS
    + PHONE_KEYS
    + CPF_KEYS
    + PLAN_KEYS
    + MONTHLY_FEE_KEYS
    + JOIN_DATE_KEYS
    + LAST_ACCESS_KEYS
    + PREFERRED_SHIFT_KEYS
    + STATUS_KEYS
    + EXTERNAL_ID_KEYS
)
_CHECKIN_PREVIEW_COLUMNS = set(
    NAME_KEYS
    + FIRST_NAME_KEYS
    + LAST_NAME_KEYS
    + EMAIL_KEYS
    + PHONE_KEYS
    + CPF_KEYS
    + PLAN_KEYS
    + EXTERNAL_ID_KEYS
    + MEMBER_ID_KEYS
    + CHECKIN_AT_KEYS
    + CHECKIN_DATE_KEYS
    + CHECKIN_TIME_KEYS
    + CHECKIN_SOURCE_KEYS
)
_MEMBER_MAPPING_TARGETS = {
    "full_name": "Nome completo",
    "first_name": "Primeiro nome",
    "last_name": "Sobrenome",
    "email": "Email",
    "phone": "Telefone",
    "cpf": "CPF",
    "plan_name": "Plano",
    "monthly_fee": "Mensalidade",
    "join_date": "Data de inicio",
    "last_checkin_at": "Ultimo acesso",
    "preferred_shift": "Turno preferido",
    "status": "Status",
    "external_id": "Matricula",
}
_CHECKIN_MAPPING_TARGETS = {
    "member_id": "ID do membro",
    "member_name": "Nome do membro",
    "first_name": "Primeiro nome",
    "last_name": "Sobrenome",
    "email": "Email",
    "external_id": "Matricula",
    "cpf": "CPF",
    "plan_name": "Plano",
    "checkin_at": "Data e hora do check-in",
    "checkin_date": "Data do check-in",
    "checkin_time": "Hora do check-in",
    "source": "Origem",
}
_MEMBER_ALIAS_TO_TARGET = {key: "full_name" for key in NAME_KEYS} | {key: "first_name" for key in FIRST_NAME_KEYS} | {
    key: "last_name" for key in LAST_NAME_KEYS
} | {key: "email" for key in EMAIL_KEYS} | {key: "phone" for key in PHONE_KEYS} | {
    key: "cpf" for key in CPF_KEYS
} | {key: "plan_name" for key in PLAN_KEYS} | {key: "monthly_fee" for key in MONTHLY_FEE_KEYS} | {
    key: "join_date" for key in JOIN_DATE_KEYS
} | {key: "last_checkin_at" for key in LAST_ACCESS_KEYS} | {
    key: "preferred_shift" for key in PREFERRED_SHIFT_KEYS
} | {key: "status" for key in STATUS_KEYS} | {key: "external_id" for key in EXTERNAL_ID_KEYS}
_CHECKIN_ALIAS_TO_TARGET = {key: "member_name" for key in NAME_KEYS} | {key: "first_name" for key in FIRST_NAME_KEYS} | {
    key: "last_name" for key in LAST_NAME_KEYS
} | {key: "email" for key in EMAIL_KEYS} | {key: "cpf" for key in CPF_KEYS} | {
    key: "plan_name" for key in PLAN_KEYS
} | {key: "external_id" for key in EXTERNAL_ID_KEYS} | {key: "member_id" for key in MEMBER_ID_KEYS} | {
    key: "checkin_at" for key in CHECKIN_AT_KEYS
} | {key: "checkin_date" for key in CHECKIN_DATE_KEYS} | {
    key: "checkin_time" for key in CHECKIN_TIME_KEYS
} | {key: "source" for key in CHECKIN_SOURCE_KEYS}


def _should_create_import_onboarding(join_date: date | None) -> bool:
    if join_date is None:
        return False
    today = datetime.now(tz=timezone.utc).date()
    if join_date > today:
        return False
    return (today - join_date).days <= _IMPORT_ONBOARDING_WINDOW_DAYS


def _build_preview_columns(seen_columns: set[str], allowed_columns: set[str]) -> tuple[list[str], list[str]]:
    recognized = sorted(column for column in seen_columns if column in allowed_columns)
    unrecognized = sorted(column for column in seen_columns if column not in allowed_columns)
    return recognized, unrecognized


def _normalize_mapping_inputs(
    column_mappings: dict[str, str] | None,
    ignored_columns: list[str] | None,
    target_options: dict[str, str],
) -> tuple[dict[str, str], set[str]]:
    normalized_mappings: dict[str, str] = {}
    for source, target in (column_mappings or {}).items():
        source_key = _normalize_header(str(source))
        target_key = _normalize_header(str(target))
        if source_key and target_key and target_key in target_options:
            normalized_mappings[source_key] = target_key

    normalized_ignored = {
        _normalize_header(str(source))
        for source in (ignored_columns or [])
        if _normalize_header(str(source))
    }
    for source_key in normalized_ignored:
        normalized_mappings.pop(source_key, None)
    return normalized_mappings, normalized_ignored


def _validate_mapping_commit(column_mappings: dict[str, str]) -> None:
    target_usage: Counter[str] = Counter(column_mappings.values())
    conflicting_targets = sorted(target for target, count in target_usage.items() if count > 1)
    if conflicting_targets:
        raise ValueError("Existem colunas mapeadas para o mesmo campo de destino.")


def _apply_column_mapping(
    row: dict[str, str],
    column_mappings: dict[str, str],
    ignored_columns: set[str],
) -> dict[str, str]:
    mapped: dict[str, str] = {}
    for source_key, value in row.items():
        if source_key in ignored_columns:
            continue
        target_key = column_mappings.get(source_key, source_key)
        if not target_key:
            continue
        if target_key not in mapped or not mapped[target_key]:
            mapped[target_key] = value
    return mapped


def _collect_source_samples(
    sample_values: dict[str, list[str]],
    row: dict[str, str],
) -> None:
    for source_key, value in row.items():
        if not value:
            continue
        bucket = sample_values.setdefault(source_key, [])
        if value in bucket:
            continue
        if len(bucket) >= 2:
            continue
        bucket.append(value)


def _humanize_source_label(source_key: str) -> str:
    return source_key.replace("_", " ").strip() or source_key


def _suggest_mapping_target(
    source_key: str,
    target_options: dict[str, str],
    alias_to_target: dict[str, str],
) -> str | None:
    direct_target = alias_to_target.get(source_key)
    if direct_target in target_options:
        return direct_target

    matches = get_close_matches(source_key, list(target_options.keys()), n=1, cutoff=0.65)
    if matches:
        return matches[0]
    return None


def _build_mapping_preview(
    *,
    seen_columns: set[str],
    sample_values: dict[str, list[str]],
    allowed_columns: set[str],
    target_options: dict[str, str],
    alias_to_target: dict[str, str],
    column_mappings: dict[str, str],
    ignored_columns: set[str],
    valid_rows: int,
) -> tuple[list[str], list[str], list[ImportPreviewSourceColumn], list[str], list[str], bool]:
    target_usage: Counter[str] = Counter(column_mappings.values())
    conflicting_targets = sorted(target for target, count in target_usage.items() if count > 1)

    source_columns: list[ImportPreviewSourceColumn] = []
    for source_key in sorted(seen_columns):
        applied_target = column_mappings.get(source_key)
        suggested_target = None
        status = "recognized"

        if source_key in ignored_columns:
            status = "ignored"
        elif applied_target and applied_target in conflicting_targets:
            status = "conflict"
        elif applied_target:
            status = "mapped"
        elif source_key in allowed_columns:
            status = "recognized"
        else:
            status = "needs_mapping"
            suggested_target = _suggest_mapping_target(source_key, target_options, alias_to_target)

        source_columns.append(
            ImportPreviewSourceColumn(
                source_key=source_key,
                source_label=_humanize_source_label(source_key),
                status=status,
                suggested_target=suggested_target,
                applied_target=applied_target,
                sample_values=sample_values.get(source_key, []),
            )
        )

    recognized_columns = sorted(
        column.source_key for column in source_columns if column.status in {"recognized", "mapped"}
    )
    unrecognized_columns = sorted(column.source_key for column in source_columns if column.status == "needs_mapping")
    blocking_issues: list[str] = []
    if conflicting_targets:
        blocking_issues.append("Existem colunas mapeadas para o mesmo campo de destino.")
    if valid_rows == 0:
        blocking_issues.append("Nenhuma linha valida foi encontrada com o mapeamento atual.")

    mapping_required = any(column.status in {"needs_mapping", "conflict"} for column in source_columns)
    can_confirm = valid_rows > 0 and not conflicting_targets
    return recognized_columns, unrecognized_columns, source_columns, conflicting_targets, blocking_issues, can_confirm


def _preview_member_snapshot(
    *,
    full_name: str,
    email: str | None,
    plan_name: str,
    join_date: date,
    row: dict[str, str],
    action: str,
) -> dict:
    return {
        "full_name": full_name,
        "email": email,
        "plan_name": plan_name,
        "join_date": join_date.isoformat(),
        "preferred_shift": _extract_preferred_shift(row),
        "action": action,
    }


def preview_members_csv(
    db: Session,
    csv_content: bytes,
    filename: str | None = None,
    *,
    column_mappings: dict[str, str] | None = None,
    ignored_columns: list[str] | None = None,
) -> ImportPreview:
    errors: list[ImportErrorEntry] = []
    warnings: list[str] = []
    sample_rows: list[ImportPreviewRow] = []
    total_rows = 0
    valid_rows = 0
    would_create = 0
    would_update = 0
    would_skip = 0
    seen_columns: set[str] = set()
    source_samples: dict[str, list[str]] = {}
    normalized_mappings, normalized_ignored = _normalize_mapping_inputs(
        column_mappings,
        ignored_columns,
        _MEMBER_MAPPING_TARGETS,
    )
    _validate_mapping_commit(normalized_mappings)

    existing_members = list(db.scalars(select(Member).where(Member.deleted_at.is_(None))).all())
    lookup = _build_member_lookups(existing_members)
    seen_emails: set[str] = set()
    seen_external_ids: set[str] = set()
    seen_cpfs: set[str] = set()
    warned_missing_join_date = False
    warned_import_task_window = False
    warned_updates = False

    for row_number, row in _iter_rows(csv_content, filename=filename):
        total_rows += 1
        seen_columns.update(row.keys())
        _collect_source_samples(source_samples, row)
        mapped_row = _apply_column_mapping(row, normalized_mappings, normalized_ignored)
        full_name = _extract_member_name(mapped_row)
        if not full_name:
            errors.append(ImportErrorEntry(row_number=row_number, reason="Nome ausente", payload=mapped_row))
            continue

        email = _truncate(((_pick_first(mapped_row, EMAIL_KEYS) or "").lower() or None), _MAX_MEMBER_EMAIL)
        external_id = _normalize_external_id(_pick_first(mapped_row, EXTERNAL_ID_KEYS))
        cpf_digits = _digits(_pick_first(mapped_row, CPF_KEYS))

        if email and email in seen_emails:
            would_skip += 1
            continue
        if external_id and external_id in seen_external_ids:
            would_skip += 1
            continue
        if cpf_digits and cpf_digits in seen_cpfs:
            would_skip += 1
            continue

        try:
            _parse_decimal(_pick_first(mapped_row, MONTHLY_FEE_KEYS))
            canonical_join_date = _parse_date(_pick_first(mapped_row, JOIN_DATE_KEYS))
            join_date = canonical_join_date or datetime.now(tz=timezone.utc).date()
            _parse_datetime(_pick_first(mapped_row, LAST_ACCESS_KEYS))
        except ValueError:
            errors.append(
                ImportErrorEntry(
                    row_number=row_number,
                    reason="Formato invalido de valor/data",
                    payload=mapped_row,
                )
            )
            continue

        valid_rows += 1
        existing_member = _resolve_member_from_row(mapped_row, lookup)
        plan_name, _, _ = _extract_plan_metadata(mapped_row, join_date=join_date)
        action = "update_member" if existing_member else "create_member"

        if existing_member:
            would_update += 1
            if not warned_updates:
                warnings.append("Linhas que batem com membros existentes serao atualizadas no commit final.")
                warned_updates = True
        else:
            would_create += 1
            if canonical_join_date is None and not warned_missing_join_date:
                warnings.append("Linhas sem data confiavel de inicio nao geram onboarding automatico apos a importacao.")
                warned_missing_join_date = True
            if (
                canonical_join_date is not None
                and _should_create_import_onboarding(canonical_join_date)
                and not warned_import_task_window
            ):
                warnings.append(
                    "Importacoes recentes geram no maximo a proxima acao operacional do onboarding/follow-up, sem retroagir o playbook completo."
                )
                warned_import_task_window = True

        if len(sample_rows) < _IMPORT_PREVIEW_SAMPLE_LIMIT:
            sample_rows.append(
                ImportPreviewRow(
                    row_number=row_number,
                    action=action,
                    preview=_preview_member_snapshot(
                        full_name=full_name,
                        email=email,
                        plan_name=plan_name,
                        join_date=join_date,
                        row=mapped_row,
                        action=action,
                    ),
                )
            )

        if email:
            seen_emails.add(email)
        if external_id:
            seen_external_ids.add(external_id)
        if cpf_digits:
            seen_cpfs.add(cpf_digits)

        if not existing_member:
            preview_member = Member(
                full_name=full_name,
                email=email,
                phone=_normalize_phone(_pick_first(mapped_row, PHONE_KEYS)),
                cpf_encrypted=encrypt_cpf(cpf_digits) if cpf_digits else None,
                status=_parse_member_status(_pick_first(mapped_row, STATUS_KEYS)),
                plan_name=plan_name,
                monthly_fee=_parse_decimal(_pick_first(mapped_row, MONTHLY_FEE_KEYS)),
                join_date=join_date,
                loyalty_months=_compute_loyalty_months(join_date),
                preferred_shift=_extract_preferred_shift(mapped_row),
                last_checkin_at=_parse_datetime(_pick_first(mapped_row, LAST_ACCESS_KEYS)),
                extra_data={},
            )
            if external_id:
                preview_member.extra_data = {"external_id": external_id}
            _add_member_to_lookups(preview_member, lookup)

    (
        recognized_columns,
        unrecognized_columns,
        source_columns,
        conflicting_targets,
        blocking_issues,
        can_confirm,
    ) = _build_mapping_preview(
        seen_columns=seen_columns,
        sample_values=source_samples,
        allowed_columns=_MEMBER_PREVIEW_COLUMNS,
        target_options=_MEMBER_MAPPING_TARGETS,
        alias_to_target=_MEMBER_ALIAS_TO_TARGET,
        column_mappings=normalized_mappings,
        ignored_columns=normalized_ignored,
        valid_rows=valid_rows,
    )
    return ImportPreview(
        preview_kind="members",
        total_rows=total_rows,
        valid_rows=valid_rows,
        would_create=would_create,
        would_update=would_update,
        would_skip=would_skip,
        recognized_columns=recognized_columns,
        unrecognized_columns=unrecognized_columns,
        mapping_required=bool(unrecognized_columns or conflicting_targets),
        can_confirm=can_confirm,
        resolved_mappings=normalized_mappings,
        ignored_columns=sorted(normalized_ignored),
        conflicting_targets=conflicting_targets,
        blocking_issues=blocking_issues,
        source_columns=source_columns,
        warnings=warnings,
        sample_rows=sample_rows,
        errors=errors,
    )


def preview_checkins_csv(
    db: Session,
    csv_content: bytes,
    filename: str | None = None,
    *,
    auto_create_missing_members: bool = False,
    column_mappings: dict[str, str] | None = None,
    ignored_columns: list[str] | None = None,
) -> ImportPreview:
    errors: list[ImportErrorEntry] = []
    warnings: list[str] = []
    sample_rows: list[ImportPreviewRow] = []
    total_rows = 0
    valid_rows = 0
    would_create = 0
    would_skip = 0
    ignored = 0
    provisional_members_possible = 0
    seen_columns: set[str] = set()
    source_samples: dict[str, list[str]] = {}
    seen_entries: set[tuple[str, str]] = set()
    pending_existing_rows: list[tuple[Member, datetime, int, dict[str, str]]] = []
    missing_member_counts: Counter[str] = Counter()
    missing_member_plans: dict[str, str | None] = {}
    normalized_mappings, normalized_ignored = _normalize_mapping_inputs(
        column_mappings,
        ignored_columns,
        _CHECKIN_MAPPING_TARGETS,
    )

    existing_members = list(db.scalars(select(Member).where(Member.deleted_at.is_(None))).all())
    lookup = _build_member_lookups(existing_members)

    for row_number, row in _iter_rows(csv_content, filename=filename):
        total_rows += 1
        seen_columns.update(row.keys())
        _collect_source_samples(source_samples, row)
        mapped_row = _apply_column_mapping(row, normalized_mappings, normalized_ignored)
        if _is_ignorable_checkin_row(mapped_row):
            ignored += 1
            continue

        parsed = _parse_checkin_datetime(
            checkin_raw=_pick_first(mapped_row, CHECKIN_AT_KEYS),
            date_raw=_pick_first(mapped_row, CHECKIN_DATE_KEYS),
            time_raw=_pick_first(mapped_row, CHECKIN_TIME_KEYS),
        )
        if not parsed:
            errors.append(ImportErrorEntry(row_number=row_number, reason="Formato de data invalido", payload=mapped_row))
            continue

        member = _resolve_member_from_row(mapped_row, lookup)
        if not member and auto_create_missing_members:
            full_name = _extract_member_name(mapped_row)
            if _is_viable_member_name(full_name):
                provisional_members_possible += 1
                valid_rows += 1
                would_create += 1
                if len(sample_rows) < _IMPORT_PREVIEW_SAMPLE_LIMIT:
                    sample_rows.append(
                        ImportPreviewRow(
                            row_number=row_number,
                            action="create_provisional_member",
                            preview={
                                "full_name": full_name,
                                "checkin_at": parsed.isoformat(),
                                "plan_name": _extract_plan_name(mapped_row),
                                "action": "create_provisional_member",
                            },
                        )
                    )
                continue

        if not member:
            missing_name = _extract_member_name(mapped_row)
            if missing_name:
                missing_member_counts[missing_name] += 1
                missing_member_plans.setdefault(missing_name, _extract_plan_name(mapped_row))
            would_skip += 1
            errors.append(
                ImportErrorEntry(
                    row_number=row_number,
                    reason="Membro nao encontrado na base de alunos importada (use member_id, email, matricula, cpf ou nome)",
                    payload=mapped_row,
                )
            )
            continue

        valid_rows += 1
        unique_key = (str(member.id), parsed.isoformat())
        if unique_key in seen_entries:
            would_skip += 1
            continue
        seen_entries.add(unique_key)
        pending_existing_rows.append((member, parsed, row_number, mapped_row))

    existing_keys = _fetch_existing_checkin_keys(db, [(member.id, parsed) for member, parsed, _, _ in pending_existing_rows])
    for member, parsed, row_number, row in pending_existing_rows:
        unique_key = (str(member.id), parsed.isoformat())
        if unique_key in existing_keys:
            would_skip += 1
            continue
        would_create += 1
        if len(sample_rows) < _IMPORT_PREVIEW_SAMPLE_LIMIT:
            sample_rows.append(
                ImportPreviewRow(
                    row_number=row_number,
                    action="create_checkin",
                    preview={
                        "member_id": str(member.id),
                        "member_name": member.full_name,
                        "checkin_at": parsed.isoformat(),
                        "action": "create_checkin",
                    },
                )
            )

    if auto_create_missing_members and provisional_members_possible > 0:
        warnings.append("Cadastros provisiorios criados por catraca nao geram onboarding automatico.")

    (
        recognized_columns,
        unrecognized_columns,
        source_columns,
        conflicting_targets,
        blocking_issues,
        can_confirm,
    ) = _build_mapping_preview(
        seen_columns=seen_columns,
        sample_values=source_samples,
        allowed_columns=_CHECKIN_PREVIEW_COLUMNS,
        target_options=_CHECKIN_MAPPING_TARGETS,
        alias_to_target=_CHECKIN_ALIAS_TO_TARGET,
        column_mappings=normalized_mappings,
        ignored_columns=normalized_ignored,
        valid_rows=valid_rows,
    )
    return ImportPreview(
        preview_kind="checkins",
        total_rows=total_rows,
        valid_rows=valid_rows,
        would_create=would_create,
        would_skip=would_skip,
        ignored_rows=ignored,
        provisional_members_possible=provisional_members_possible,
        recognized_columns=recognized_columns,
        unrecognized_columns=unrecognized_columns,
        mapping_required=bool(unrecognized_columns or conflicting_targets),
        can_confirm=can_confirm,
        resolved_mappings=normalized_mappings,
        ignored_columns=sorted(normalized_ignored),
        conflicting_targets=conflicting_targets,
        blocking_issues=blocking_issues,
        source_columns=source_columns,
        missing_members=_build_missing_member_entries(missing_member_counts, missing_member_plans),
        warnings=warnings,
        sample_rows=sample_rows,
        errors=errors,
    )


def import_members_csv(
    db: Session,
    csv_content: bytes,
    filename: str | None = None,
    *,
    column_mappings: dict[str, str] | None = None,
    ignored_columns: list[str] | None = None,
) -> ImportSummary:
    errors: list[ImportErrorEntry] = []
    duplicates = 0
    imported = 0
    updated_existing = 0
    pending_count = 0
    touched_members: list[Member] = []
    normalized_mappings, normalized_ignored = _normalize_mapping_inputs(
        column_mappings,
        ignored_columns,
        _MEMBER_MAPPING_TARGETS,
    )
    _validate_mapping_commit(normalized_mappings)

    existing_members = list(db.scalars(select(Member).where(Member.deleted_at.is_(None))).all())
    lookup = _build_member_lookups(existing_members)
    seen_emails: set[str] = set()
    seen_external_ids: set[str] = set()
    seen_cpfs: set[str] = set()

    for row_number, row in _iter_rows(csv_content, filename=filename):
        mapped_row = _apply_column_mapping(row, normalized_mappings, normalized_ignored)
        full_name = _extract_member_name(mapped_row)
        if not full_name:
            errors.append(ImportErrorEntry(row_number=row_number, reason="Nome ausente", payload=mapped_row))
            continue

        email = _truncate(((_pick_first(mapped_row, EMAIL_KEYS) or "").lower() or None), _MAX_MEMBER_EMAIL)
        external_id = _normalize_external_id(_pick_first(mapped_row, EXTERNAL_ID_KEYS))
        cpf_digits = _digits(_pick_first(mapped_row, CPF_KEYS))

        if email and email in seen_emails:
            duplicates += 1
            continue
        if external_id and external_id in seen_external_ids:
            duplicates += 1
            continue
        if cpf_digits and cpf_digits in seen_cpfs:
            duplicates += 1
            continue

        monthly_fee_raw = _pick_first(mapped_row, MONTHLY_FEE_KEYS)
        join_date_raw = _pick_first(mapped_row, JOIN_DATE_KEYS)
        try:
            monthly_fee = _parse_decimal(monthly_fee_raw)
            canonical_join_date = _parse_date(join_date_raw)
            join_date = canonical_join_date or datetime.now(tz=timezone.utc).date()
            last_checkin_at = _parse_datetime(_pick_first(mapped_row, LAST_ACCESS_KEYS))
        except ValueError:
            errors.append(
                ImportErrorEntry(
                    row_number=row_number,
                    reason="Formato invalido de valor/data",
                    payload=mapped_row,
                )
            )
            continue

        existing_member = _resolve_member_from_row(mapped_row, lookup)
        if existing_member:
            if _refresh_existing_member_from_import_row(
                existing_member,
                mapped_row,
                email=email,
                external_id=external_id,
                cpf_digits=cpf_digits,
                monthly_fee=monthly_fee,
                join_date=join_date,
                last_checkin_at=last_checkin_at,
            ):
                db.add(existing_member)
                pending_count += 1
                touched_members.append(existing_member)
            _add_member_to_lookups(existing_member, lookup)
            updated_existing += 1
            if email:
                seen_emails.add(email)
            if external_id:
                seen_external_ids.add(external_id)
            if cpf_digits:
                seen_cpfs.add(cpf_digits)
            if pending_count >= _IMPORT_BATCH_SIZE:
                db.commit()
                pending_count = 0
            continue

        plan_name, plan_cycle, plan_cycle_source = _extract_plan_metadata(mapped_row, join_date=join_date)
        extra_data: dict = {"imported": True}
        if external_id:
            extra_data["external_id"] = external_id
        _populate_member_extra_data(extra_data, mapped_row)
        extra_data["plan_cycle"] = plan_cycle or _PLAN_CYCLE_SOURCE_UNKNOWN
        extra_data["plan_cycle_source"] = plan_cycle_source

        member = Member(
            full_name=full_name,
            email=email,
            phone=_normalize_phone(_pick_first(mapped_row, PHONE_KEYS)),
            cpf_encrypted=encrypt_cpf(cpf_digits) if cpf_digits else None,
            status=_parse_member_status(_pick_first(mapped_row, STATUS_KEYS)),
            plan_name=plan_name,
            monthly_fee=monthly_fee,
            join_date=join_date,
            loyalty_months=_compute_loyalty_months(join_date),
            preferred_shift=_extract_preferred_shift(mapped_row),
            last_checkin_at=last_checkin_at,
            extra_data=extra_data,
        )
        db.add(member)
        if _should_create_import_onboarding(canonical_join_date):
            db.flush()
            create_import_playbook_tasks_for_member(db, member, commit=False)
        imported += 1
        pending_count += 1
        touched_members.append(member)
        _add_member_to_lookups(member, lookup)

        if email:
            seen_emails.add(email)
        if external_id:
            seen_external_ids.add(external_id)
        if cpf_digits:
            seen_cpfs.add(cpf_digits)

        if pending_count >= _IMPORT_BATCH_SIZE:
            db.commit()
            pending_count = 0

    if pending_count:
        db.commit()
    touched_member_ids = [member.id for member in touched_members if member.id]
    if touched_member_ids:
        refresh_member_risk_snapshot(db, member_ids=touched_member_ids, sync_alerts=True)
        db.commit()
    if imported or updated_existing:
        invalidate_dashboard_cache("members", "risk")
    return ImportSummary(
        imported=imported,
        updated_existing=updated_existing,
        skipped_duplicates=duplicates,
        ignored_rows=0,
        errors=errors,
    )


def import_checkins_csv(
    db: Session,
    csv_content: bytes,
    filename: str | None = None,
    *,
    auto_create_missing_members: bool = False,
    column_mappings: dict[str, str] | None = None,
    ignored_columns: list[str] | None = None,
) -> ImportSummary:
    errors: list[ImportErrorEntry] = []
    duplicates = 0
    ignored = 0
    imported = 0
    provisional_created = 0
    provisional_members: list[str] = []
    seen_entries: set[tuple[str, str]] = set()
    pending_rows: list[tuple[Member, datetime, CheckinSource, dict[str, str]]] = []
    touched_member_ids: set[UUID] = set()
    missing_member_counts: Counter[str] = Counter()
    missing_member_plans: dict[str, str | None] = {}
    normalized_mappings, normalized_ignored = _normalize_mapping_inputs(
        column_mappings,
        ignored_columns,
        _CHECKIN_MAPPING_TARGETS,
    )
    _validate_mapping_commit(normalized_mappings)

    existing_members = list(db.scalars(select(Member).where(Member.deleted_at.is_(None))).all())
    lookup = _build_member_lookups(existing_members)

    for row_number, row in _iter_rows(csv_content, filename=filename):
        mapped_row = _apply_column_mapping(row, normalized_mappings, normalized_ignored)
        if _is_ignorable_checkin_row(mapped_row):
            ignored += 1
            continue

        date_raw = _pick_first(mapped_row, CHECKIN_DATE_KEYS)
        time_raw = _pick_first(mapped_row, CHECKIN_TIME_KEYS)
        checkin_raw = _pick_first(mapped_row, CHECKIN_AT_KEYS)
        parsed = _parse_checkin_datetime(checkin_raw=checkin_raw, date_raw=date_raw, time_raw=time_raw)
        if not parsed:
            errors.append(ImportErrorEntry(row_number=row_number, reason="Formato de data invalido", payload=mapped_row))
            continue

        member = _resolve_member_from_row(mapped_row, lookup)
        if not member and auto_create_missing_members:
            member = _create_provisional_member_from_checkin(db, mapped_row, parsed)
            if member:
                provisional_created += 1
                provisional_members.append(member.full_name)
                _add_member_to_lookups(member, lookup)

        if not member:
            missing_name = _extract_member_name(mapped_row)
            if missing_name:
                missing_member_counts[missing_name] += 1
                missing_member_plans.setdefault(missing_name, _extract_plan_name(mapped_row))
            errors.append(
                ImportErrorEntry(
                    row_number=row_number,
                    reason="Membro nao encontrado na base de alunos importada (use member_id, email, matricula, cpf ou nome)",
                    payload=mapped_row,
                )
            )
            continue

        unique_key = (str(member.id), parsed.isoformat())
        if unique_key in seen_entries:
            duplicates += 1
            continue
        seen_entries.add(unique_key)
        pending_rows.append(
            (
                member,
                parsed,
                _parse_checkin_source(_pick_first(mapped_row, CHECKIN_SOURCE_KEYS)),
                mapped_row,
            )
        )

    existing_keys = _fetch_existing_checkin_keys(db, [(member.id, parsed) for member, parsed, _, _ in pending_rows])

    for member, parsed, source, row in pending_rows:
        unique_key = (str(member.id), parsed.isoformat())
        if unique_key in existing_keys:
            duplicates += 1
            continue

        checkin = Checkin(
            member_id=member.id,
            checkin_at=parsed,
            source=source,
            hour_bucket=parsed.hour,
            weekday=parsed.weekday(),
            extra_data={"imported": True, "raw": row},
        )
        if member.last_checkin_at is None or parsed > member.last_checkin_at:
            member.last_checkin_at = parsed
            db.add(member)
        touched_member_ids.add(member.id)
        db.add(checkin)
        imported += 1

    db.commit()
    if touched_member_ids:
        refresh_member_risk_snapshot(db, member_ids=touched_member_ids, sync_alerts=True)
        db.commit()
    if touched_member_ids:
        invalidate_dashboard_cache("checkins", "risk")
    if provisional_created:
        invalidate_dashboard_cache("members")
    return ImportSummary(
        imported=imported,
        skipped_duplicates=duplicates,
        ignored_rows=ignored,
        provisional_members_created=provisional_created,
        provisional_members=provisional_members,
        missing_members=_build_missing_member_entries(missing_member_counts, missing_member_plans),
        errors=errors,
    )


def _fetch_existing_checkin_keys(db: Session, keys: list[tuple[UUID, datetime]]) -> set[tuple[str, str]]:
    if not keys:
        return set()

    existing_keys: set[tuple[str, str]] = set()
    chunk_size = 1000
    for idx in range(0, len(keys), chunk_size):
        chunk = keys[idx : idx + chunk_size]
        rows = db.execute(
            select(Checkin.member_id, Checkin.checkin_at).where(tuple_(Checkin.member_id, Checkin.checkin_at).in_(chunk))
        ).all()
        for member_id, checkin_at in rows:
            if checkin_at.tzinfo is None:
                normalized = checkin_at.replace(tzinfo=timezone.utc)
            else:
                normalized = checkin_at.astimezone(timezone.utc)
            existing_keys.add((str(member_id), normalized.isoformat()))
    return existing_keys


def _decode_csv_text(csv_content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return csv_content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return csv_content.decode("utf-8", errors="ignore")


def _iter_rows(csv_content: bytes, filename: str | None = None):
    kind = _detect_file_kind(csv_content, filename)
    if kind == "xlsx":
        yield from _iter_rows_xlsx(csv_content)
        return
    yield from _iter_rows_csv(csv_content)


def _detect_file_kind(content: bytes, filename: str | None) -> str:
    if filename:
        lower = filename.lower()
        if lower.endswith(".xlsx"):
            return "xlsx"
        if lower.endswith(".csv"):
            return "csv"
    if content.startswith(b"PK\x03\x04"):
        return "xlsx"
    return "csv"


def _iter_rows_csv(csv_content: bytes):
    text = _decode_csv_text(csv_content)
    delimiter = _detect_delimiter(text)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    for row_number, row in enumerate(reader, start=2):
        yield row_number, _normalize_row(row)


def _iter_rows_xlsx(content: bytes):
    try:
        from openpyxl import load_workbook
    except ImportError:
        yield from _iter_rows_xlsx_fallback(content)
        return

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Arquivo XLSX invalido ou corrompido.") from exc

    try:
        worksheet = workbook.active
        header_cells = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_cells:
            return

        headers = [str(cell).strip() if cell is not None else "" for cell in header_cells]
        for row_number, row_cells in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            raw_row: dict[str | None, str | None] = {}
            has_value = False
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = row_cells[idx] if idx < len(row_cells) else None
                normalized_value = _xlsx_cell_to_text(value)
                if normalized_value:
                    has_value = True
                raw_row[header] = normalized_value
            if not has_value:
                continue
            yield row_number, _normalize_row(raw_row)
    finally:
        workbook.close()


def _iter_rows_xlsx_fallback(content: bytes):
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            worksheet_path = _xlsx_first_sheet_path(archive)
            shared_strings = _xlsx_shared_strings(archive)
            worksheet_xml = archive.read(worksheet_path)
    except Exception as exc:
        raise ValueError("Arquivo XLSX invalido ou corrompido.") from exc

    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    root = ET.fromstring(worksheet_xml)
    rows = root.findall(".//x:sheetData/x:row", ns)
    if not rows:
        return

    header_values = _xlsx_row_values(rows[0], shared_strings, ns)
    headers = [str(value).strip() if value is not None else "" for value in header_values]
    for row in rows[1:]:
        row_number = int(row.attrib.get("r", "0") or "0")
        values = _xlsx_row_values(row, shared_strings, ns)
        raw_row: dict[str | None, str | None] = {}
        has_value = False
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = values[idx] if idx < len(values) else None
            normalized_value = _xlsx_cell_to_text(value)
            if normalized_value:
                has_value = True
            raw_row[header] = normalized_value
        if not has_value:
            continue
        yield (row_number or 0), _normalize_row(raw_row)


def _xlsx_first_sheet_path(archive: zipfile.ZipFile) -> str:
    default_path = "xl/worksheets/sheet1.xml"
    if default_path in archive.namelist():
        return default_path

    workbook_xml = archive.read("xl/workbook.xml")
    workbook_rels_xml = archive.read("xl/_rels/workbook.xml.rels")

    workbook_ns = {
        "x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    rels_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}

    workbook_root = ET.fromstring(workbook_xml)
    first_sheet = workbook_root.find(".//x:sheets/x:sheet", workbook_ns)
    if first_sheet is None:
        raise ValueError("Planilha XLSX sem abas.")
    rel_id = first_sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
    if not rel_id:
        raise ValueError("Planilha XLSX sem relacao de aba.")

    rels_root = ET.fromstring(workbook_rels_xml)
    for rel in rels_root.findall(".//r:Relationship", rels_ns):
        if rel.attrib.get("Id") != rel_id:
            continue
        target = rel.attrib.get("Target", "").lstrip("/")
        if not target:
            break
        if target.startswith("xl/"):
            return target
        return f"xl/{target}"
    raise ValueError("Nao foi possivel localizar a aba principal no XLSX.")


def _xlsx_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    path = "xl/sharedStrings.xml"
    if path not in archive.namelist():
        return []

    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    root = ET.fromstring(archive.read(path))
    values: list[str] = []
    for si in root.findall(".//x:si", ns):
        texts = [node.text or "" for node in si.findall(".//x:t", ns)]
        values.append("".join(texts))
    return values


def _xlsx_row_values(row: ET.Element, shared_strings: list[str], ns: dict[str, str]) -> list[object | None]:
    values: list[object | None] = []
    for cell in row.findall("x:c", ns):
        ref = cell.attrib.get("r", "")
        col_index = _xlsx_col_index(ref)
        while len(values) <= col_index:
            values.append(None)
        values[col_index] = _xlsx_cell_value(cell, shared_strings, ns)
    return values


def _xlsx_col_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha()).upper()
    if not letters:
        return 0
    index = 0
    for letter in letters:
        index = index * 26 + (ord(letter) - ord("A") + 1)
    return index - 1


def _xlsx_cell_value(cell: ET.Element, shared_strings: list[str], ns: dict[str, str]) -> object | None:
    cell_type = cell.attrib.get("t", "")
    if cell_type == "inlineStr":
        node = cell.find("x:is/x:t", ns)
        return node.text if node is not None else ""

    value_node = cell.find("x:v", ns)
    raw_value = value_node.text if value_node is not None else ""
    if raw_value is None:
        return ""

    if cell_type == "s":
        try:
            return shared_strings[int(raw_value)]
        except Exception:
            return raw_value
    if cell_type == "b":
        return "1" if raw_value == "1" else "0"
    return raw_value


def _xlsx_cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _detect_delimiter(text: str) -> str:
    sample = "\n".join(text.splitlines()[:10]) or text
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
        return dialect.delimiter
    except csv.Error:
        first_line = sample.splitlines()[0] if sample else ""
        return ";" if first_line.count(";") > first_line.count(",") else ","


def _normalize_row(row: dict[str | None, str | None]) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for key, value in row.items():
        if not key:
            continue
        normalized_key = _normalize_header(key)
        if not normalized_key:
            continue
        normalized[normalized_key] = (value or "").strip()
    return normalized


def _normalize_header(value: str) -> str:
    collapsed = _normalize_text(value)
    normalized = re.sub(r"[^a-z0-9]+", "_", collapsed).strip("_")
    return normalized


def _normalize_text(value: str) -> str:
    ascii_text = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", ascii_text).strip().lower()


def _pick_first(row: dict[str, str], aliases: tuple[str, ...]) -> str | None:
    for key in aliases:
        value = row.get(key)
        if value:
            return value.strip()
    return None


def _truncate(value: str | None, max_length: int) -> str | None:
    if value is None:
        return None
    trimmed = re.sub(r"\s+", " ", value).strip()
    if not trimmed:
        return None
    if len(trimmed) <= max_length:
        return trimmed
    return trimmed[:max_length]


def _normalize_phone(raw_value: str | None) -> str | None:
    if not raw_value:
        return None
    text_value = raw_value.strip()
    if not text_value:
        return None

    digits_only = _digits(text_value)
    if digits_only:
        # Keep the first BR-style phone chunk when multiple numbers are concatenated.
        if len(digits_only) >= 11:
            return digits_only[:11]
        if len(digits_only) >= 10:
            return digits_only[:10]
        return digits_only[:_MAX_MEMBER_PHONE]

    candidates = re.findall(r"\d{8,}", text_value)
    if candidates:
        return candidates[0][:_MAX_MEMBER_PHONE]
    return _truncate(text_value, _MAX_MEMBER_PHONE)


def _digits(value: str | None) -> str:
    return re.sub(r"\D+", "", value or "")


def _parse_decimal(value: str | None) -> Decimal:
    if not value:
        return Decimal("0")
    cleaned = value.strip().replace("R$", "").replace(" ", "")
    if not cleaned:
        return Decimal("0")
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    try:
        return Decimal(cleaned)
    except InvalidOperation as exc:
        raise ValueError("decimal invalido") from exc


def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    raw = value.strip()
    if not raw:
        return None

    serial = _parse_excel_serial(raw)
    if serial is not None:
        return _excel_serial_to_datetime(serial).date()

    if "T" in raw:
        raw = raw.split("T", 1)[0]
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(raw).date()
    except ValueError:
        return None


def _parse_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    raw = value.strip()
    if not raw:
        return None

    serial = _parse_excel_serial(raw)
    if serial is not None:
        return _excel_serial_to_datetime(serial).replace(tzinfo=timezone.utc)

    iso_candidate = raw.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(iso_candidate)
    except ValueError:
        parsed = None

    if parsed is None:
        for fmt in DATETIME_FORMATS:
            try:
                parsed = datetime.strptime(raw, fmt)
                break
            except ValueError:
                continue

    if parsed is None:
        parsed_date = _parse_date(raw)
        if parsed_date:
            parsed = datetime.combine(parsed_date, time.min)
        else:
            return None

    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _parse_time(value: str | None) -> time | None:
    if not value:
        return None
    raw = value.strip()
    if not raw:
        return None

    serial = _parse_excel_serial(raw)
    if serial is not None and 0 <= serial < 1:
        return _excel_serial_to_datetime(serial).time()

    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(raw, fmt).time()
        except ValueError:
            continue
    return None


def _parse_checkin_datetime(*, checkin_raw: str | None, date_raw: str | None, time_raw: str | None) -> datetime | None:
    if date_raw and time_raw:
        parsed_date = _parse_date(date_raw)
        parsed_time = _parse_time(time_raw)
        if parsed_date and parsed_time:
            return datetime.combine(parsed_date, parsed_time, tzinfo=timezone.utc)

    candidate = checkin_raw or date_raw
    parsed = _parse_datetime(candidate)
    if parsed:
        return parsed

    if date_raw and time_raw:
        return _parse_datetime(f"{date_raw} {time_raw}")
    return None


def _parse_member_status(raw_value: str | None) -> MemberStatus:
    key = _normalize_header(raw_value or "")
    mapping = {
        "active": MemberStatus.ACTIVE,
        "ativo": MemberStatus.ACTIVE,
        "a": MemberStatus.ACTIVE,
        "paused": MemberStatus.PAUSED,
        "pausado": MemberStatus.PAUSED,
        "p": MemberStatus.PAUSED,
        "cancelled": MemberStatus.CANCELLED,
        "canceled": MemberStatus.CANCELLED,
        "cancelado": MemberStatus.CANCELLED,
        "inactive": MemberStatus.CANCELLED,
        "inativo": MemberStatus.CANCELLED,
        "i": MemberStatus.CANCELLED,
    }
    return mapping.get(key, MemberStatus.ACTIVE)


_EXCEL_SERIAL_MIN = 35000  # ~1995-10-09
_EXCEL_SERIAL_MAX = 55000  # ~2050-07-03


def _parse_excel_serial(raw_value: str) -> float | None:
    cleaned = raw_value.strip().replace(",", ".")
    if not cleaned:
        return None
    # Must be a plain number — if it contains date/time separators it's not an Excel serial.
    if any(ch in cleaned for ch in ("-", "/", ":", "T")):
        return None
    try:
        serial = float(cleaned)
    except ValueError:
        return None
    if 0 <= serial < 1:
        return serial
    # Only accept serials within a plausible date range (~1995 to 2050).
    if not (_EXCEL_SERIAL_MIN <= serial <= _EXCEL_SERIAL_MAX):
        return None
    return serial


def _excel_serial_to_datetime(serial: float) -> datetime:
    # Excel uses 1899-12-30 as base in the 1900 date system.
    base = datetime(1899, 12, 30)
    days = int(serial)
    fraction = serial - days
    seconds = int(round(fraction * 86400))
    if seconds >= 86400:
        days += 1
        seconds -= 86400
    return base + timedelta(days=days, seconds=seconds)


def _parse_checkin_source(raw_value: str | None) -> CheckinSource:
    key = _normalize_header(raw_value or "")
    mapping = {
        "manual": CheckinSource.MANUAL,
        "turnstile": CheckinSource.TURNSTILE,
        "catraca": CheckinSource.TURNSTILE,
        "import": CheckinSource.IMPORT,
        "importado": CheckinSource.IMPORT,
    }
    return mapping.get(key, CheckinSource.IMPORT)


def _build_member_lookups(members: list[Member]) -> dict[str, dict]:
    by_id: dict[str, Member] = {}
    by_email: dict[str, Member] = {}
    by_external_id: dict[str, Member] = {}
    by_cpf: dict[str, Member] = {}
    by_name: dict[str, list[Member]] = {}
    by_name_compact: dict[str, list[Member]] = {}
    by_name_core: dict[str, list[Member]] = {}

    for member in members:
        by_id[str(member.id)] = member

        if member.email:
            by_email[member.email.lower()] = member

        extra_data = member.extra_data or {}
        for external_id in _external_id_candidates(str(extra_data.get("external_id") or "")):
            by_external_id[external_id] = member

        if member.cpf_encrypted:
            try:
                cpf_digits = _digits(decrypt_cpf(member.cpf_encrypted))
                if cpf_digits and cpf_digits not in by_cpf:
                    by_cpf[cpf_digits] = member
            except Exception:
                pass

        name_key = _normalize_text(member.full_name)
        by_name.setdefault(name_key, []).append(member)
        compact_key = _compact_name(member.full_name)
        if compact_key:
            by_name_compact.setdefault(compact_key, []).append(member)
        core_key = _core_name_key(member.full_name)
        if core_key:
            by_name_core.setdefault(core_key, []).append(member)

    return {
        "by_id": by_id,
        "by_email": by_email,
        "by_external_id": by_external_id,
        "by_cpf": by_cpf,
        "by_name": by_name,
        "by_name_compact": by_name_compact,
        "by_name_core": by_name_core,
    }


def _add_member_to_lookups(member: Member, lookup: dict[str, dict]) -> None:
    lookup["by_id"][str(member.id)] = member

    if member.email:
        lookup["by_email"][member.email.lower()] = member

    extra_data = member.extra_data or {}
    for external_id in _external_id_candidates(str(extra_data.get("external_id") or "")):
        lookup["by_external_id"][external_id] = member

    if member.cpf_encrypted:
        try:
            cpf_digits = _digits(decrypt_cpf(member.cpf_encrypted))
            if cpf_digits and cpf_digits not in lookup["by_cpf"]:
                lookup["by_cpf"][cpf_digits] = member
        except Exception:
            pass

    name_key = _normalize_text(member.full_name)
    _append_lookup_member(lookup["by_name"], name_key, member)
    compact_key = _compact_name(member.full_name)
    if compact_key:
        _append_lookup_member(lookup["by_name_compact"], compact_key, member)
    core_key = _core_name_key(member.full_name)
    if core_key:
        _append_lookup_member(lookup["by_name_core"], core_key, member)


def _append_lookup_member(bucket: dict[str, list[Member]], key: str, member: Member) -> None:
    candidates = bucket.setdefault(key, [])
    member_id = getattr(member, "id", None)
    for existing in candidates:
        if existing is member:
            return
        if member_id is not None and getattr(existing, "id", None) == member_id:
            return
    candidates.append(member)


def _find_existing_member_by_import_keys(
    lookup: dict[str, dict],
    *,
    email: str | None,
    external_id: str | None,
    cpf_digits: str | None,
) -> Member | None:
    if email:
        member = lookup["by_email"].get(email)
        if member:
            return member
    if external_id:
        for candidate in _external_id_candidates(external_id):
            member = lookup["by_external_id"].get(candidate)
            if member:
                return member
    if cpf_digits:
        member = lookup["by_cpf"].get(cpf_digits)
        if member:
            return member
    return None


def _refresh_existing_member_from_import_row(
    member: Member,
    row: dict[str, str],
    *,
    email: str | None,
    external_id: str | None,
    cpf_digits: str | None,
    monthly_fee: Decimal,
    join_date: date,
    last_checkin_at: datetime | None,
) -> bool:
    changed = False
    extra_data = dict(member.extra_data or {})

    if email and member.email != email:
        member.email = email
        changed = True
    phone = _normalize_phone(_pick_first(row, PHONE_KEYS))
    if phone and member.phone != phone:
        member.phone = phone
        changed = True
    if cpf_digits and not member.cpf_encrypted:
        member.cpf_encrypted = encrypt_cpf(cpf_digits)
        changed = True

    status = _parse_member_status(_pick_first(row, STATUS_KEYS))
    if member.status != status:
        member.status = status
        changed = True

    plan_name, plan_cycle, plan_cycle_source = _extract_plan_metadata(row, join_date=join_date)
    if member.plan_name != plan_name:
        member.plan_name = plan_name
        changed = True
    if member.monthly_fee != monthly_fee:
        member.monthly_fee = monthly_fee
        changed = True
    if member.join_date != join_date:
        member.join_date = join_date
        member.loyalty_months = _compute_loyalty_months(join_date)
        changed = True

    preferred_shift = _extract_preferred_shift(row)
    if member.preferred_shift != preferred_shift:
        member.preferred_shift = preferred_shift
        changed = True
    if last_checkin_at and (member.last_checkin_at is None or last_checkin_at > member.last_checkin_at):
        member.last_checkin_at = last_checkin_at
        changed = True

    if external_id:
        previous_external_id = str(extra_data.get("external_id") or "")
        if previous_external_id != external_id:
            extra_data["external_id"] = external_id
            changed = True

    before_snapshot = dict(extra_data)
    _populate_member_extra_data(extra_data, row)
    extra_data["plan_cycle"] = plan_cycle or _PLAN_CYCLE_SOURCE_UNKNOWN
    extra_data["plan_cycle_source"] = plan_cycle_source
    if extra_data != before_snapshot:
        changed = True

    if changed:
        member.extra_data = extra_data
    return changed


def _resolve_member_from_row(row: dict[str, str], lookup: dict[str, dict]) -> Member | None:
    member_id_raw = _pick_first(row, MEMBER_ID_KEYS)
    if member_id_raw:
        try:
            parsed_uuid = UUID(member_id_raw)
            member = lookup["by_id"].get(str(parsed_uuid))
            if member:
                return member
        except ValueError:
            # Segue para outras estrategias de match.
            pass

    email = (_pick_first(row, EMAIL_KEYS) or "").lower()
    if email:
        member = lookup["by_email"].get(email)
        if member:
            return member

    external_id = _normalize_external_id(_pick_first(row, EXTERNAL_ID_KEYS) or member_id_raw)
    for candidate in _external_id_candidates(external_id):
        member = lookup["by_external_id"].get(candidate)
        if member:
            return member

    cpf_digits = _digits(_pick_first(row, CPF_KEYS))
    if cpf_digits:
        member = lookup["by_cpf"].get(cpf_digits)
        if member:
            return member

    name = _extract_member_name(row)
    if name:
        candidates = lookup["by_name"].get(_normalize_text(name), [])
        best_candidate = _select_best_member_candidate(candidates, row)
        if best_candidate:
            return best_candidate
        candidates = lookup["by_name_compact"].get(_compact_name(name), [])
        best_candidate = _select_best_member_candidate(candidates, row)
        if best_candidate:
            return best_candidate
        candidates = lookup["by_name_core"].get(_core_name_key(name), [])
        best_candidate = _select_best_member_candidate(candidates, row)
        if best_candidate:
            return best_candidate
        fuzzy = _find_unique_name_candidate(name, lookup["by_name_compact"])
        if fuzzy:
            return fuzzy

    return None


def _extract_member_name(row: dict[str, str]) -> str | None:
    direct = _truncate(_pick_first(row, NAME_KEYS), _MAX_MEMBER_NAME)
    if direct:
        return direct
    first_name = _pick_first(row, FIRST_NAME_KEYS)
    last_name = _pick_first(row, LAST_NAME_KEYS)
    combined = " ".join(part for part in [first_name, last_name] if part)
    return _truncate(combined, _MAX_MEMBER_NAME)


def _is_viable_member_name(name: str | None) -> bool:
    if not name:
        return False
    normalized = _normalize_text(name)
    if len(normalized) < 5:
        return False
    if normalized in _IGNORABLE_CHECKIN_NAMES or normalized.startswith("total de registros"):
        return False
    return bool(re.search(r"[a-z]", normalized))


def _create_provisional_member_from_checkin(db: Session, row: dict[str, str], parsed: datetime) -> Member | None:
    full_name = _extract_member_name(row)
    if not _is_viable_member_name(full_name):
        return None

    cpf_digits = _digits(_pick_first(row, CPF_KEYS))
    extra_data: dict = {
        "imported": True,
        "provisional_member": True,
        "provisional_source": "checkin_import",
        "provisional_created_at": datetime.now(tz=timezone.utc).isoformat(),
    }
    external_id = _normalize_external_id(_pick_first(row, EXTERNAL_ID_KEYS))
    if external_id:
        extra_data["external_id"] = external_id
    _populate_member_extra_data(extra_data, row)

    provisional_join_date = parsed.date()
    plan_name, plan_cycle, plan_cycle_source = _extract_plan_metadata(row, join_date=provisional_join_date)
    extra_data["plan_cycle"] = plan_cycle or _PLAN_CYCLE_SOURCE_UNKNOWN
    extra_data["plan_cycle_source"] = plan_cycle_source
    member = Member(
        full_name=full_name,
        email=_truncate(((_pick_first(row, EMAIL_KEYS) or "").lower() or None), _MAX_MEMBER_EMAIL),
        phone=_normalize_phone(_pick_first(row, PHONE_KEYS)),
        cpf_encrypted=encrypt_cpf(cpf_digits) if cpf_digits else None,
        status=MemberStatus.ACTIVE,
        plan_name=plan_name,
        monthly_fee=Decimal("0"),
        join_date=provisional_join_date,
        loyalty_months=_compute_loyalty_months(provisional_join_date),
        preferred_shift=_extract_preferred_shift(row),
        last_checkin_at=parsed,
        extra_data=extra_data,
    )
    db.add(member)
    db.flush()
    return member


def _is_ignorable_checkin_row(row: dict[str, str]) -> bool:
    raw_name = _extract_member_name(row)
    if not raw_name:
        return False
    normalized = _normalize_text(raw_name)
    if normalized in _IGNORABLE_CHECKIN_NAMES:
        return True
    return normalized.startswith("total de registros")


def _build_missing_member_entries(counts: Counter[str], plans: dict[str, str | None]) -> list[MissingMemberEntry]:
    entries = [
        MissingMemberEntry(name=name, occurrences=occurrences, sample_plan=plans.get(name))
        for name, occurrences in counts.most_common()
    ]
    return entries


def _extract_plan_name(row: dict[str, str]) -> str:
    plan_name, _, _ = _extract_plan_metadata(row)
    return plan_name


def _extract_plan_metadata(row: dict[str, str], join_date: date | None = None) -> tuple[str, str | None, str]:
    primary_label = _pick_first(row, PLAN_NAME_SOURCE_KEYS)
    conditions_label = _pick_first(row, PLAN_CONDITION_KEYS)
    renewal_date = _parse_date(_pick_first(row, PLAN_RENEWAL_KEYS))
    effective_join_date = join_date or _parse_date(_pick_first(row, JOIN_DATE_KEYS))

    cycle = _infer_plan_cycle_from_text(conditions_label)
    cycle_source = "conditions"
    if not cycle:
        cycle = _infer_plan_cycle_from_dates(effective_join_date, renewal_date)
        cycle_source = "dates"
    if not cycle:
        cycle = _infer_plan_cycle_from_text(primary_label)
        cycle_source = "plan_name"
    if not cycle:
        cycle_source = _PLAN_CYCLE_SOURCE_UNKNOWN

    base_label = _extract_plan_base_label(primary_label, conditions_label, cycle)
    if cycle and base_label:
        return _compose_plan_name(base_label, cycle), cycle, cycle_source

    for candidate in (primary_label, conditions_label, _pick_first(row, PLAN_KEYS)):
        trimmed = _truncate(candidate, _MAX_MEMBER_PLAN)
        if not trimmed:
            continue
        normalized = _normalize_text(trimmed)
        if normalized in _GENERIC_PLAN_VALUES and not cycle:
            continue
        if cycle:
            cleaned = _clean_plan_base_label(trimmed)
            if cleaned:
                return _compose_plan_name(cleaned, cycle), cycle, cycle_source
        return trimmed, cycle, cycle_source
    return "Plano Base", cycle, cycle_source


def _infer_plan_cycle_from_text(raw_value: str | None) -> str | None:
    if not raw_value:
        return None
    normalized = _normalize_text(raw_value)
    if not normalized:
        return None
    for cycle, pattern in _PLAN_CYCLE_PATTERNS.items():
        if pattern.search(normalized):
            return cycle
    return None


def _infer_plan_cycle_from_dates(join_date: date | None, renewal_date: date | None) -> str | None:
    if join_date is None or renewal_date is None:
        return None
    delta_days = (renewal_date - join_date).days
    if 330 <= delta_days <= 380:
        return "annual"
    if 150 <= delta_days <= 210:
        return "semiannual"
    if 25 <= delta_days <= 45:
        return "monthly"
    return None


def _extract_plan_base_label(primary_label: str | None, conditions_label: str | None, cycle: str | None) -> str | None:
    for candidate in (primary_label, conditions_label):
        if not candidate:
            continue
        normalized = _normalize_text(candidate)
        if normalized in _GENERIC_PLAN_VALUES and cycle is None:
            continue
        cleaned = _clean_plan_base_label(candidate)
        if cleaned:
            return cleaned
        trimmed = _truncate(candidate, _MAX_MEMBER_PLAN)
        if trimmed:
            return trimmed
    return None


def _clean_plan_base_label(raw_value: str | None) -> str | None:
    if not raw_value:
        return None
    cleaned = _PLAN_CYCLE_CLEANUP_PATTERN.sub("", raw_value)
    cleaned = re.sub(r"\(\s*\)", "", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned)
    parts = []
    seen_parts: set[str] = set()
    for part in cleaned.split(","):
        normalized_part = re.sub(r"\s+", " ", part).strip(" -/(),")
        if not normalized_part:
            continue
        dedupe_key = normalized_part.lower()
        if dedupe_key in seen_parts:
            continue
        seen_parts.add(dedupe_key)
        parts.append(normalized_part)
    cleaned = (parts[0] if parts else "").strip(" -/(),")
    return _truncate(cleaned, _MAX_MEMBER_PLAN)


def _compose_plan_name(base_label: str, cycle: str) -> str:
    cycle_label = _PLAN_CYCLE_LABELS[cycle]
    base = _truncate(base_label, _MAX_MEMBER_PLAN)
    if not base:
        return cycle_label.title()
    if cycle_label in base.upper():
        return base
    return _truncate(f"{base} {cycle_label}", _MAX_MEMBER_PLAN) or base


def _candidate_plan_match(member: Member, row: dict[str, str]) -> bool:
    raw_plan = _pick_first(row, PLAN_KEYS)
    if not raw_plan:
        return False
    normalized_plan = _normalize_text(raw_plan)
    if not normalized_plan or normalized_plan in _GENERIC_PLAN_VALUES:
        return False

    member_values = [
        member.plan_name,
        str((member.extra_data or {}).get("raw_plan_name") or ""),
        str((member.extra_data or {}).get("raw_plan_conditions") or ""),
    ]
    return any(_normalize_text(value) == normalized_plan for value in member_values if value)


def _member_candidate_rank(member: Member, row: dict[str, str]) -> tuple:
    status_rank = {
        MemberStatus.ACTIVE: 3,
        MemberStatus.PAUSED: 2,
        MemberStatus.CANCELLED: 1,
    }.get(member.status, 0)
    return (
        status_rank,
        1 if _candidate_plan_match(member, row) else 0,
        1 if member.email else 0,
        1 if member.phone else 0,
        1 if member.cpf_encrypted else 0,
        1 if member.last_checkin_at else 0,
        member.join_date.toordinal() if member.join_date else 0,
        int(member.updated_at.timestamp()) if member.updated_at else 0,
        str(member.id),
    )


def _select_best_member_candidate(candidates: list[Member], row: dict[str, str]) -> Member | None:
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]
    return max(candidates, key=lambda member: _member_candidate_rank(member, row))


def _extract_preferred_shift(row: dict[str, str]) -> str | None:
    candidate = _truncate(_pick_first(row, PREFERRED_SHIFT_KEYS), _MAX_MEMBER_SHIFT)
    if not candidate:
        return None
    normalized = _normalize_text(candidate)
    if normalized in _GENERIC_SHIFT_VALUES:
        return None
    return candidate


def refresh_member_plan_metadata(db: Session, gym_id=None) -> int:
    filters = [Member.deleted_at.is_(None)]
    if gym_id is not None:
        filters.append(Member.gym_id == gym_id)

    members = list(db.scalars(select(Member).where(*filters)).all())
    updated = 0
    for member in members:
        row = _build_plan_row_from_member(member)
        plan_name, plan_cycle, plan_cycle_source = _extract_plan_metadata(row, join_date=member.join_date)
        normalized_cycle = plan_cycle or _PLAN_CYCLE_SOURCE_UNKNOWN
        extra_data = dict(member.extra_data or {})

        if (
            member.plan_name == plan_name
            and extra_data.get("plan_cycle") == normalized_cycle
            and extra_data.get("plan_cycle_source") == plan_cycle_source
        ):
            continue

        extra_data["plan_cycle"] = normalized_cycle
        extra_data["plan_cycle_source"] = plan_cycle_source
        if not extra_data.get("raw_plan_name") and member.plan_name:
            extra_data["raw_plan_name"] = _truncate(member.plan_name, 255)
        member.plan_name = plan_name
        member.extra_data = extra_data
        db.add(member)
        updated += 1

    if updated:
        db.commit()
        invalidate_dashboard_cache("members")
    return updated


def _build_plan_row_from_member(member: Member) -> dict[str, str]:
    extra_data = member.extra_data or {}
    return {
        "plan_name": member.plan_name,
        "assinaturas": str(extra_data.get("raw_plan_name") or ""),
        "assinaturas_condicoes": str(extra_data.get("raw_plan_conditions") or ""),
        "dt_prox_renovacao_assinatura": str(extra_data.get("next_plan_renewal_raw") or ""),
        "data_prox_vencimento": str(extra_data.get("next_due_date_raw") or ""),
    }


def _populate_member_extra_data(extra_data: dict, row: dict[str, str]) -> None:
    metadata_map = {
        "cidade_residencial": "city",
        "estado_residencial": "state",
        "sexo": "gender",
        "tipo_cadastro": "registration_type",
        "categoria": "category",
        "consultores": "consultants",
        "aniversario": "birthday_label",
        "data_prox_vencimento": "next_due_date_raw",
        "dt_prox_renovacao_assinatura": "next_plan_renewal_raw",
        "dt_ultimo_recebimento": "last_payment_received_raw",
        "dt_primeira_ativacao": "first_activation_raw",
        "cadastro": "registration_date_raw",
        "assinaturas": "raw_plan_name",
        "assinatura": "raw_plan_name",
        "assinaturas_condicoes": "raw_plan_conditions",
        "assinaturas_horarios": "raw_shift_label",
    }
    for source_key, target_key in metadata_map.items():
        value = _truncate(row.get(source_key), 255)
        if value:
            extra_data[target_key] = value


def _compute_loyalty_months(join_date: date) -> int:
    delta = relativedelta(date.today(), join_date)
    return max(0, delta.years * 12 + delta.months)


def _normalize_external_id(raw_value: str | None) -> str | None:
    if not raw_value:
        return None
    value = raw_value.strip().lower()
    return value or None


def _external_id_candidates(raw_value: str | None) -> tuple[str, ...]:
    normalized = _normalize_external_id(raw_value)
    if not normalized:
        return ()
    candidates: list[str] = [normalized]
    digits = _digits(normalized)
    if digits:
        candidates.append(digits)
        stripped = digits.lstrip("0")
        if stripped:
            candidates.append(stripped)
    return tuple(dict.fromkeys(candidate for candidate in candidates if candidate))


def _compact_name(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", "", _normalize_text(value or ""))


def _core_name_key(value: str | None) -> str:
    tokens = [token for token in _normalize_text(value or "").split(" ") if token and token not in _NAME_PARTICLES]
    return " ".join(tokens)


def _find_unique_name_candidate(name: str, lookup: dict[str, list[Member]]) -> Member | None:
    target = _compact_name(name)
    if len(target) < 8:
        return None
    matches: dict[UUID, Member] = {}
    for key, members in lookup.items():
        if not key:
            continue
        if target in key or key in target:
            for member in members:
                matches[member.id] = member
    if len(matches) == 1:
        return next(iter(matches.values()))
    return None
